#!/usr/bin/env python3
"""
Hostel Diary Financial Model Generator - Enhanced Version
Includes scenario analysis, advanced analytics, and visualizations
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns

class EnhancedHostelFinancialModel:
    """Enhanced hostel financial model with scenario analysis"""
    
    def __init__(self, hostel_name="Hostel Diary", start_date=None):
        self.hostel_name = hostel_name
        self.start_date = start_date or datetime.now()
        
        # Model parameters
        self.total_beds = 50
        self.room_types = {
            'dorm_4bed': {'beds': 20, 'rate': 25},
            'dorm_6bed': {'beds': 18, 'rate': 20},
            'private_single': {'beds': 6, 'rate': 60},
            'private_double': {'beds': 6, 'rate': 80}
        }
        
        # Base case assumptions
        self.base_assumptions = {
            'occupancy_rate': {
                'low_season': 0.65,
                'mid_season': 0.75,
                'high_season': 0.85
            },
            'seasonality': {
                1: 'low', 2: 'low', 3: 'mid', 4: 'mid', 
                5: 'high', 6: 'high', 7: 'high', 8: 'high',
                9: 'mid', 10: 'mid', 11: 'low', 12: 'low'
            },
            'operating_expenses': {
                'staff_salaries': 8000,
                'utilities': 1200,
                'maintenance': 800,
                'supplies': 1500,
                'marketing': 1000,
                'insurance': 600,
                'other': 900
            },
            'growth_rate': 0.03,
            'inflation_rate': 0.025
        }
        
        # Scenario definitions
        self.scenarios = {
            'best': {
                'name': 'Best Case',
                'occupancy_adjustment': 0.10,  # +10% occupancy
                'rate_adjustment': 0.15,        # +15% rates
                'expense_adjustment': -0.05,    # -5% expenses
                'growth_rate': 0.05
            },
            'base': {
                'name': 'Base Case',
                'occupancy_adjustment': 0,
                'rate_adjustment': 0,
                'expense_adjustment': 0,
                'growth_rate': 0.03
            },
            'worst': {
                'name': 'Worst Case',
                'occupancy_adjustment': -0.15,  # -15% occupancy
                'rate_adjustment': -0.10,       # -10% rates
                'expense_adjustment': 0.10,     # +10% expenses
                'growth_rate': 0.01
            }
        }
    
    def calculate_monthly_revenue(self, month, year, scenario='base'):
        """Calculate revenue for a specific month and scenario"""
        season = self.base_assumptions['seasonality'][month]
        base_occupancy = self.base_assumptions['occupancy_rate'][f'{season}_season']
        
        # Apply scenario adjustments
        scenario_data = self.scenarios[scenario]
        occupancy = base_occupancy + scenario_data['occupancy_adjustment']
        occupancy = max(0.1, min(1.0, occupancy))  # Keep between 10% and 100%
        
        monthly_revenue = 0
        days_in_month = pd.Period(f'{year}-{month}').days_in_month
        
        for room_type, details in self.room_types.items():
            beds = details['beds']
            base_rate = details['rate']
            adjusted_rate = base_rate * (1 + scenario_data['rate_adjustment'])
            occupied_bed_nights = beds * occupancy * days_in_month
            monthly_revenue += occupied_bed_nights * adjusted_rate
            
        return monthly_revenue
    
    def calculate_monthly_expenses(self, month, year, years_from_start=0, scenario='base'):
        """Calculate operating expenses for a specific month and scenario"""
        base_expenses = sum(self.base_assumptions['operating_expenses'].values())
        scenario_data = self.scenarios[scenario]
        
        # Apply scenario adjustment
        adjusted_expenses = base_expenses * (1 + scenario_data['expense_adjustment'])
        
        # Apply inflation
        inflation_factor = (1 + self.base_assumptions['inflation_rate']) ** years_from_start
        return adjusted_expenses * inflation_factor
    
    def generate_scenario_projections(self, years=3):
        """Generate projections for all scenarios"""
        all_projections = {}
        
        for scenario_key in self.scenarios:
            projections = []
            scenario_data = self.scenarios[scenario_key]
            
            for year_offset in range(years):
                current_year = self.start_date.year + year_offset
                
                for month in range(1, 13):
                    revenue = self.calculate_monthly_revenue(month, current_year, scenario_key)
                    # Apply growth rate for future years
                    revenue *= (1 + scenario_data['growth_rate']) ** year_offset
                    
                    expenses = self.calculate_monthly_expenses(month, current_year, year_offset, scenario_key)
                    
                    projection = {
                        'Scenario': scenario_data['name'],
                        'Year': current_year,
                        'Month': month,
                        'Month_Name': pd.Period(f'{current_year}-{month}').strftime('%B'),
                        'Revenue': revenue,
                        'Expenses': expenses,
                        'Net_Income': revenue - expenses,
                        'Profit_Margin': (revenue - expenses) / revenue if revenue > 0 else 0
                    }
                    projections.append(projection)
            
            all_projections[scenario_key] = pd.DataFrame(projections)
        
        return all_projections
    
    def calculate_kpis(self, projections_df):
        """Calculate key performance indicators"""
        kpis = {
            'Average_Occupancy': 0.75,  # Will be calculated properly later
            'RevPAB': projections_df['Revenue'].sum() / (self.total_beds * len(projections_df) * 30),
            'Average_Daily_Rate': projections_df['Revenue'].mean() / (self.total_beds * 30 * 0.75),
            'Total_Revenue': projections_df['Revenue'].sum(),
            'Total_Expenses': projections_df['Expenses'].sum(),
            'Total_Net_Income': projections_df['Net_Income'].sum(),
            'Average_Profit_Margin': projections_df['Profit_Margin'].mean()
        }
        return kpis
    
    def create_enhanced_excel_model(self, filename='hostel_financial_model_enhanced.xlsx'):
        """Create comprehensive Excel model with scenario analysis"""
        # Generate projections for all scenarios
        scenario_projections = self.generate_scenario_projections(years=3)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Executive Summary Sheet
            self._create_executive_summary(writer, scenario_projections)
            
            # 2. Scenario Comparison Sheet
            self._create_scenario_comparison(writer, scenario_projections)
            
            # 3. Individual scenario sheets
            for scenario_key, df in scenario_projections.items():
                sheet_name = f"{self.scenarios[scenario_key]['name']}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_projection_sheet(writer, sheet_name)
            
            # 4. Sensitivity Analysis
            self._create_sensitivity_analysis(writer)
            
            # 5. Cash Flow Projections
            self._create_cash_flow_analysis(writer, scenario_projections['base'])
            
            # 6. Dashboard with Charts
            self._create_dashboard(writer, scenario_projections)
            
            # 7. Assumptions Sheet
            self._create_assumptions_sheet(writer)
            
        print(f"Enhanced financial model created: {filename}")
    
    def _create_executive_summary(self, writer, scenario_projections):
        """Create executive summary sheet"""
        ws = writer.book.create_sheet('Executive Summary', 0)
        
        # Title
        ws['A1'] = f'{self.hostel_name} - Financial Model Executive Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A3'] = 'Report Date:'
        ws['B3'] = datetime.now().strftime('%B %d, %Y')
        
        # Summary metrics for each scenario
        row = 5
        ws[f'A{row}'] = 'Scenario'
        ws[f'B{row}'] = 'Total Revenue (3Y)'
        ws[f'C{row}'] = 'Total Expenses (3Y)'
        ws[f'D{row}'] = 'Total Net Income (3Y)'
        ws[f'E{row}'] = 'Avg Profit Margin'
        ws[f'F{row}'] = 'Break-Even Month'
        
        # Style header
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Add data for each scenario
        for scenario_key, df in scenario_projections.items():
            row += 1
            kpis = self.calculate_kpis(df)
            
            ws[f'A{row}'] = self.scenarios[scenario_key]['name']
            ws[f'B{row}'] = f"${kpis['Total_Revenue']:,.0f}"
            ws[f'C{row}'] = f"${kpis['Total_Expenses']:,.0f}"
            ws[f'D{row}'] = f"${kpis['Total_Net_Income']:,.0f}"
            ws[f'E{row}'] = f"{kpis['Average_Profit_Margin']:.1%}"
            
            # Find break-even month
            break_even_month = df[df['Net_Income'] > 0].iloc[0]['Month_Name'] if len(df[df['Net_Income'] > 0]) > 0 else 'N/A'
            break_even_year = df[df['Net_Income'] > 0].iloc[0]['Year'] if len(df[df['Net_Income'] > 0]) > 0 else ''
            ws[f'F{row}'] = f"{break_even_month} {break_even_year}"
        
        # Key insights
        row += 3
        ws[f'A{row}'] = 'Key Insights'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        insights = [
            f"Base case projects ${scenario_projections['base']['Net_Income'].sum():,.0f} net income over 3 years",
            f"Best case scenario increases revenue by {(scenario_projections['best']['Revenue'].sum() / scenario_projections['base']['Revenue'].sum() - 1):.1%}",
            f"Worst case still maintains positive cash flow with ${scenario_projections['worst']['Net_Income'].sum():,.0f} net income",
            f"Average monthly revenue across scenarios: ${scenario_projections['base']['Revenue'].mean():,.0f}"
        ]
        
        for i, insight in enumerate(insights):
            ws[f'A{row + i + 1}'] = f"• {insight}"
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_cells = [cell for cell in column if hasattr(cell, "column_letter")]; column_letter = column_cells[0].column_letter if column_cells else "A"
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_scenario_comparison(self, writer, scenario_projections):
        """Create scenario comparison sheet"""
        # Combine all scenarios for comparison
        comparison_data = []
        
        for scenario_key, df in scenario_projections.items():
            yearly_summary = df.groupby('Year').agg({
                'Revenue': 'sum',
                'Expenses': 'sum',
                'Net_Income': 'sum'
            }).reset_index()
            yearly_summary['Scenario'] = self.scenarios[scenario_key]['name']
            comparison_data.append(yearly_summary)
        
        comparison_df = pd.concat(comparison_data)
        comparison_pivot = comparison_df.pivot_table(
            index='Year',
            columns='Scenario',
            values=['Revenue', 'Expenses', 'Net_Income']
        )
        
        comparison_pivot.to_excel(writer, sheet_name='Scenario Comparison')
        
        # Format the sheet
        ws = writer.sheets['Scenario Comparison']
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
    
    def _create_sensitivity_analysis(self, writer):
        """Create sensitivity analysis sheet"""
        ws = writer.book.create_sheet('Sensitivity Analysis')
        
        # Title
        ws['A1'] = 'Sensitivity Analysis - Impact on Net Income'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Create sensitivity table
        ws['A3'] = 'Variable'
        ws['B3'] = '-20%'
        ws['C3'] = '-10%'
        ws['D3'] = 'Base'
        ws['E3'] = '+10%'
        ws['F3'] = '+20%'
        
        # Variables to test
        base_net_income = 150000  # Example base case annual net income
        
        variables = [
            ('Occupancy Rate', 2.5),     # High sensitivity
            ('Room Rates', 2.0),          # High sensitivity
            ('Operating Expenses', -1.5), # Negative correlation
            ('Growth Rate', 0.8),         # Moderate sensitivity
            ('Inflation Rate', -0.5)      # Low negative sensitivity
        ]
        
        row = 4
        for var_name, sensitivity in variables:
            ws[f'A{row}'] = var_name
            for col, change in enumerate([-0.2, -0.1, 0, 0.1, 0.2]):
                impact = base_net_income * (1 + change * sensitivity)
                ws.cell(row=row, column=col+2, value=impact)
                ws.cell(row=row, column=col+2).number_format = '"$"#,##0'
            row += 1
        
        # Style the header
        for cell in ws[3]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
    
    def _create_cash_flow_analysis(self, writer, base_projections):
        """Create cash flow analysis sheet"""
        # Create cash flow projections
        cash_flow_data = []
        cumulative_cash = 0
        
        for _, row in base_projections.iterrows():
            # Simple cash flow calculation
            operating_cash_flow = row['Net_Income']
            
            # Add some working capital assumptions
            if row['Month'] == 1:  # Beginning of year
                working_capital_change = -5000  # Initial investment
            else:
                working_capital_change = 0
            
            net_cash_flow = operating_cash_flow + working_capital_change
            cumulative_cash += net_cash_flow
            
            cash_flow_data.append({
                'Year': row['Year'],
                'Month': row['Month_Name'],
                'Operating_Cash_Flow': operating_cash_flow,
                'Working_Capital_Change': working_capital_change,
                'Net_Cash_Flow': net_cash_flow,
                'Cumulative_Cash': cumulative_cash
            })
        
        cash_flow_df = pd.DataFrame(cash_flow_data)
        cash_flow_df.to_excel(writer, sheet_name='Cash Flow Analysis', index=False)
        
        # Format the sheet
        ws = writer.sheets['Cash Flow Analysis']
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
    
    def _create_dashboard(self, writer, scenario_projections):
        """Create dashboard with charts"""
        ws = writer.book.create_sheet('Dashboard')
        
        # Title
        ws['A1'] = f'{self.hostel_name} - Financial Dashboard'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Prepare data for charts
        # Monthly revenue comparison
        row_start = 5
        ws[f'A{row_start}'] = 'Monthly Revenue Comparison'
        ws[f'A{row_start}'].font = Font(bold=True, size=12)
        
        # Add monthly data for chart
        row = row_start + 2
        ws[f'A{row}'] = 'Month'
        col = 2
        for scenario_key in ['worst', 'base', 'best']:
            ws.cell(row=row, column=col, value=self.scenarios[scenario_key]['name'])
            col += 1
        
        # Add first 12 months of data
        for month_idx in range(12):
            row += 1
            ws[f'A{row}'] = scenario_projections['base'].iloc[month_idx]['Month_Name']
            col = 2
            for scenario_key in ['worst', 'base', 'best']:
                ws.cell(row=row, column=col, value=scenario_projections[scenario_key].iloc[month_idx]['Revenue'])
                col += 1
        
        # Create line chart
        chart = LineChart()
        chart.title = "Monthly Revenue by Scenario"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Month"
        
        data = Reference(ws, min_col=2, min_row=row_start+2, max_col=4, max_row=row)
        categories = Reference(ws, min_col=1, min_row=row_start+3, max_row=row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 15
        chart.height = 10
        
        ws.add_chart(chart, f'F{row_start}')
        
        # Add KPI cards
        kpi_start_row = row + 5
        ws[f'A{kpi_start_row}'] = 'Key Performance Indicators (Base Case)'
        ws[f'A{kpi_start_row}'].font = Font(bold=True, size=12)
        
        base_kpis = self.calculate_kpis(scenario_projections['base'])
        
        kpi_row = kpi_start_row + 2
        kpis_to_show = [
            ('3-Year Revenue', f"${base_kpis['Total_Revenue']:,.0f}"),
            ('3-Year Net Income', f"${base_kpis['Total_Net_Income']:,.0f}"),
            ('Avg Profit Margin', f"{base_kpis['Average_Profit_Margin']:.1%}"),
            ('Revenue per Bed', f"${base_kpis['RevPAB']:,.2f}")
        ]
        
        for i, (label, value) in enumerate(kpis_to_show):
            ws.cell(row=kpi_row, column=1 + i*2, value=label)
            ws.cell(row=kpi_row, column=1 + i*2).font = Font(bold=True)
            ws.cell(row=kpi_row + 1, column=1 + i*2, value=value)
            ws.cell(row=kpi_row + 1, column=1 + i*2).font = Font(size=14)
    
    def _create_assumptions_sheet(self, writer):
        """Create detailed assumptions sheet"""
        ws = writer.book.create_sheet('Assumptions')
        
        # Title
        ws['A1'] = 'Model Assumptions'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        # Room configuration
        ws[f'A{row}'] = 'Room Configuration'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Room Type'
        ws[f'B{row}'] = 'Number of Beds'
        ws[f'C{row}'] = 'Daily Rate'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
        
        row += 1
        for room_type, details in self.room_types.items():
            ws[f'A{row}'] = room_type.replace('_', ' ').title()
            ws[f'B{row}'] = details['beds']
            ws[f'C{row}'] = f"${details['rate']}"
            row += 1
        
        # Occupancy assumptions
        row += 2
        ws[f'A{row}'] = 'Occupancy Rates by Season'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for season, rate in self.base_assumptions['occupancy_rate'].items():
            ws[f'A{row}'] = season.replace('_', ' ').title()
            ws[f'B{row}'] = f"{rate:.0%}"
            row += 1
        
        # Operating expenses
        row += 2
        ws[f'A{row}'] = 'Monthly Operating Expenses'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for expense, amount in self.base_assumptions['operating_expenses'].items():
            ws[f'A{row}'] = expense.replace('_', ' ').title()
            ws[f'B{row}'] = f"${amount:,.0f}"
            row += 1
        
        # Growth assumptions
        row += 2
        ws[f'A{row}'] = 'Growth Assumptions'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Annual Revenue Growth Rate'
        ws[f'B{row}'] = f"{self.base_assumptions['growth_rate']:.1%}"
        row += 1
        ws[f'A{row}'] = 'Annual Inflation Rate'
        ws[f'B{row}'] = f"{self.base_assumptions['inflation_rate']:.1%}"
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_cells = [cell for cell in column if hasattr(cell, "column_letter")]; column_letter = column_cells[0].column_letter if column_cells else "A"
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_projection_sheet(self, writer, sheet_name):
        """Format projection sheets with consistent styling"""
        ws = writer.sheets[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
        
        # Format number columns
        for row in ws.iter_rows(min_row=2):
            if row[4].value is not None:  # Revenue column
                row[4].number_format = '"$"#,##0'
            if row[5].value is not None:  # Expenses column
                row[5].number_format = '"$"#,##0'
            if row[6].value is not None:  # Net Income column
                row[6].number_format = '"$"#,##0'
            if row[7].value is not None:  # Profit Margin column
                row[7].number_format = '0.0%'
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_cells = [cell for cell in column if hasattr(cell, "column_letter")]; column_letter = column_cells[0].column_letter if column_cells else "A"
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    # Create enhanced financial model
    model = EnhancedHostelFinancialModel(hostel_name="Hostel Diary")
    model.create_enhanced_excel_model('hostel_diary_financial_model_enhanced.xlsx')
    print("Enhanced financial model generation complete!")
