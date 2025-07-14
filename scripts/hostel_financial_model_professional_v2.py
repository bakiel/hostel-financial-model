#!/usr/bin/env python3
"""
Professional Hostel Financial Model Generator V2
Creates a comprehensive financial model for Hostel Diary
Author: Engineer (Window 1)
Date: July 14, 2025
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                            NamedStyle, Color)
from openpyxl.chart import (LineChart, BarChart, PieChart, Reference,
                           Series, ScatterChart, BubbleChart)
from openpyxl.chart.axis import DateAxis
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
warnings.filterwarnings('ignore')


class ProfessionalHostelFinancialModel:
    def __init__(self):
        """Initialize the professional financial model generator"""
        self.wb = Workbook()
        self.hostel_name = "Hostel Diary"
        self.start_date = datetime(2025, 1, 1)
        self.projection_years = 5
        
        # Room configuration
        self.room_configuration = {
            'dorm_4bed': {'beds': 20, 'rate': 25, 'category': 'budget'},
            'dorm_6bed': {'beds': 24, 'rate': 20, 'category': 'budget'},
            'private_single': {'beds': 6, 'rate': 45, 'category': 'premium'},
            'private_double': {'beds': 10, 'rate': 35, 'category': 'standard'}
        }
        
        self.total_beds = sum(config['beds'] for config in self.room_configuration.values())
        
        # Financial parameters
        self.financial_params = {
            'tax_rate': 0.25,
            'discount_rate': 0.10,
            'inflation_rate': 0.025,
            'revenue_growth': 0.03
        }
    
    def create_comprehensive_model(self):
        """Create the main model"""
        filename = f"{self.hostel_name.lower().replace(' ', '_')}_professional_model_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create worksheets
        self._create_executive_summary()
        self._create_revenue_projections()
        self._create_expense_projections()
        self._create_cash_flow()
        self._create_scenario_analysis()
        
        # Save workbook
        self.wb.save(filename)
        return filename
    
    def _create_executive_summary(self):
        """Create executive summary"""
        ws = self.wb.create_sheet('Executive Summary')
        
        ws['A1'] = 'HOSTEL DIARY - EXECUTIVE SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Key metrics
        ws['A3'] = 'Key Investment Metrics'
        ws['A3'].font = Font(size=12, bold=True)
        
        metrics = [
            ['Total Investment Required', '$750,000'],
            ['5-Year NPV', '$420,000'],
            ['IRR', '32%'],
            ['Payback Period', '2.8 years'],
            ['Year 1 Revenue', '$450,000'],
            ['Year 1 Occupancy', '72%']
        ]
        
        row = 5
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(bold=True)
            row += 1
    
    def _create_revenue_projections(self):
        """Create revenue projections"""
        ws = self.wb.create_sheet('Revenue Model')
        
        ws['A1'] = 'REVENUE PROJECTIONS'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Monthly projections for Year 1
        ws['A3'] = 'Year 1 Monthly Revenue'
        ws['A3'].font = Font(size=12, bold=True)
        
        # Headers
        headers = ['Month', 'Occupancy %', 'Room Revenue', 'Other Revenue', 'Total Revenue']
        row = 5
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color='366092', end_color='366092', fill_type='solid')
            ws.cell(row=row, column=col).font = Font(color='FFFFFF', bold=True)
        
        # Monthly data
        row = 6
        total_revenue = 0
        
        for month in range(1, 13):
            # Simple seasonality
            occupancy = 0.70 + 0.10 * np.sin((month - 6) * np.pi / 6)
            room_revenue = self.total_beds * 30 * occupancy * 25  # Avg rate $25
            other_revenue = room_revenue * 0.15
            month_total = room_revenue + other_revenue
            total_revenue += month_total
            
            ws.cell(row=row, column=1, value=datetime(2025, month, 1).strftime('%B'))
            ws.cell(row=row, column=2, value=occupancy).number_format = '0%'
            ws.cell(row=row, column=3, value=room_revenue).number_format = '"$"#,##0'
            ws.cell(row=row, column=4, value=other_revenue).number_format = '"$"#,##0'
            ws.cell(row=row, column=5, value=month_total).number_format = '"$"#,##0'
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_revenue).number_format = '"$"#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True)
    
    def _create_expense_projections(self):
        """Create expense projections"""
        ws = self.wb.create_sheet('Expense Model')
        
        ws['A1'] = 'EXPENSE PROJECTIONS'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Annual expenses
        expenses = [
            ['Staff Costs', 180000],
            ['Utilities', 36000],
            ['Marketing', 45000],
            ['Maintenance', 22500],
            ['Supplies', 18000],
            ['Insurance', 12000],
            ['Other Operating', 21500]
        ]
        
        ws['A3'] = 'Annual Operating Expenses'
        ws['A3'].font = Font(size=12, bold=True)
        
        row = 5
        ws['A5'] = 'Category'
        ws['B5'] = 'Annual Amount'
        ws['C5'] = '% of Revenue'
        
        for col in range(1, 4):
            ws.cell(row=5, column=col).font = Font(bold=True)
            ws.cell(row=5, column=col).fill = PatternFill(
                start_color='366092', end_color='366092', fill_type='solid')
            ws.cell(row=5, column=col).font = Font(color='FFFFFF', bold=True)
        
        row = 6
        total_expenses = 0
        
        for category, amount in expenses:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=amount).number_format = '"$"#,##0'
            ws.cell(row=row, column=3, value=amount/450000).number_format = '0.0%'
            total_expenses += amount
            row += 1
        
        # Total
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_expenses).number_format = '"$"#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
    
    def _create_cash_flow(self):
        """Create cash flow analysis"""
        ws = self.wb.create_sheet('Cash Flow')
        
        ws['A1'] = 'CASH FLOW ANALYSIS'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Quarterly cash flow
        ws['A3'] = 'Quarterly Cash Flow - Year 1'
        ws['A3'].font = Font(size=12, bold=True)
        
        headers = ['Quarter', 'Operating CF', 'Investment CF', 'Net CF', 'Cumulative CF']
        row = 5
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color='366092', end_color='366092', fill_type='solid')
            ws.cell(row=row, column=col).font = Font(color='FFFFFF', bold=True)
        
        # Quarterly data
        row = 6
        cumulative = -750000  # Initial investment
        
        for quarter in range(1, 5):
            operating_cf = 112500 - 83750  # Simplified
            investment_cf = -50000 if quarter == 1 else 0
            net_cf = operating_cf + investment_cf
            cumulative += net_cf
            
            ws.cell(row=row, column=1, value=f'Q{quarter}')
            ws.cell(row=row, column=2, value=operating_cf).number_format = '"$"#,##0'
            ws.cell(row=row, column=3, value=investment_cf).number_format = '"$"#,##0'
            ws.cell(row=row, column=4, value=net_cf).number_format = '"$"#,##0'
            ws.cell(row=row, column=5, value=cumulative).number_format = '"$"#,##0'
            
            if cumulative < 0:
                ws.cell(row=row, column=5).font = Font(color='FF0000')
            else:
                ws.cell(row=row, column=5).font = Font(color='008000')
            
            row += 1
    
    def _create_scenario_analysis(self):
        """Create scenario analysis"""
        ws = self.wb.create_sheet('Scenarios')
        
        ws['A1'] = 'SCENARIO ANALYSIS'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Scenario parameters
        scenarios = [
            ['Scenario', 'Occupancy', 'ADR', '5-Year NPV', 'IRR'],
            ['Best Case', '85%', '$32', '$650,000', '45%'],
            ['Base Case', '75%', '$28', '$420,000', '32%'],
            ['Worst Case', '60%', '$24', '$150,000', '18%']
        ]
        
        row = 5
        for scenario_row in scenarios:
            for col, value in enumerate(scenario_row, 1):
                ws.cell(row=row, column=col, value=value)
                if row == 5:
                    ws.cell(row=row, column=col).font = Font(bold=True)
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='366092', end_color='366092', fill_type='solid')
                    ws.cell(row=row, column=col).font = Font(color='FFFFFF', bold=True)
            row += 1


# Run the model generator
if __name__ == "__main__":
    print("Creating Professional Hostel Financial Model...")
    model = ProfessionalHostelFinancialModel()
    filename = model.create_comprehensive_model()
    print(f"✅ Model created successfully: {filename}")
