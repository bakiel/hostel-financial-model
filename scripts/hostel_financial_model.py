#!/usr/bin/env python3
"""
Hostel Diary Financial Model Generator
Analyzes and enhances hostel financial data with automated calculations and projections
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns

class HostelFinancialModel:
    """Main class for hostel financial modeling and analysis"""
    
    def __init__(self, hostel_name="Hostel Diary", start_date=None):
        self.hostel_name = hostel_name
        self.start_date = start_date or datetime.now()
        
        # Model parameters
        self.total_beds = 50  # Default, can be updated
        self.room_types = {
            'dorm_4bed': {'beds': 20, 'rate': 25},
            'dorm_6bed': {'beds': 18, 'rate': 20},
            'private_single': {'beds': 6, 'rate': 60},
            'private_double': {'beds': 6, 'rate': 80}
        }
        
        # Financial assumptions
        self.assumptions = {
            'occupancy_rate': {
                'low_season': 0.65,
                'mid_season': 0.75,
                'high_season': 0.85
            },
            'seasonality': {
                1: 'low', 2: 'low', 3: 'mid', 4: 'mid', 
                5: 'high', 6: 'high', 7: 'high', 8: 'high',
                9: 'mid', 10: 'mid', 11: 'low', 12: 'low'
            },
            'operating_expenses': {
                'staff_salaries': 8000,  # Monthly
                'utilities': 1200,
                'maintenance': 800,
                'supplies': 1500,
                'marketing': 1000,
                'insurance': 600,
                'other': 900
            },
            'growth_rate': 0.03,  # Annual growth
            'inflation_rate': 0.025  # Annual inflation
        }
    
    def calculate_monthly_revenue(self, month, year):
        """Calculate revenue for a specific month"""
        season = self.assumptions['seasonality'][month]
        occupancy = self.assumptions['occupancy_rate'][f'{season}_season']
        
        monthly_revenue = 0
        days_in_month = pd.Period(f'{year}-{month}').days_in_month
        
        for room_type, details in self.room_types.items():
            beds = details['beds']
            rate = details['rate']
            occupied_bed_nights = beds * occupancy * days_in_month
            monthly_revenue += occupied_bed_nights * rate
            
        return monthly_revenue
    
    def calculate_monthly_expenses(self, month, year, years_from_start=0):
        """Calculate operating expenses for a specific month"""
        base_expenses = sum(self.assumptions['operating_expenses'].values())
        inflation_factor = (1 + self.assumptions['inflation_rate']) ** years_from_start
        return base_expenses * inflation_factor
    
    def generate_projections(self, years=3):
        """Generate financial projections for specified number of years"""
        projections = []
        
        for year_offset in range(years):
            current_year = self.start_date.year + year_offset
            
            for month in range(1, 13):
                revenue = self.calculate_monthly_revenue(month, current_year)
                # Apply growth rate for future years
                revenue *= (1 + self.assumptions['growth_rate']) ** year_offset
                
                expenses = self.calculate_monthly_expenses(month, current_year, year_offset)
                
                projection = {
                    'Year': current_year,
                    'Month': month,
                    'Month_Name': pd.Period(f'{current_year}-{month}').strftime('%B'),
                    'Revenue': revenue,
                    'Expenses': expenses,
                    'Net_Income': revenue - expenses,
                    'Occupancy_Rate': self.get_occupancy_rate(month)
                }
                projections.append(projection)
                
        return pd.DataFrame(projections)
    
    def get_occupancy_rate(self, month):
        """Get occupancy rate for a specific month"""
        season = self.assumptions['seasonality'][month]
        return self.assumptions['occupancy_rate'][f'{season}_season']
    
    def calculate_break_even(self):
        """Calculate break-even occupancy rate"""
        monthly_expenses = sum(self.assumptions['operating_expenses'].values())
        
        # Calculate average daily rate
        total_revenue_capacity = sum(
            details['beds'] * details['rate'] 
            for details in self.room_types.values()
        )
        
        # Break-even occupancy
        days_per_month = 30  # Average
        total_bed_nights = self.total_beds * days_per_month
        break_even_occupancy = monthly_expenses / (total_revenue_capacity * days_per_month)
        
        return break_even_occupancy
    
    def create_excel_model(self, filename='hostel_financial_model.xlsx'):
        """Create comprehensive Excel financial model"""
        # Generate projections
        df_projections = self.generate_projections(years=3)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write projections
            df_projections.to_excel(writer, sheet_name='Projections', index=False)
            
            # Create summary sheet
            self._create_summary_sheet(writer, df_projections)
            
            # Create assumptions sheet
            self._create_assumptions_sheet(writer)
            
            # Create dashboard sheet
            self._create_dashboard_sheet(writer, df_projections)
            
            # Format sheets
            self._format_excel_sheets(writer)
            
        print(f"Financial model created: {filename}")
        return filename
    
    def _create_summary_sheet(self, writer, projections_df):
        """Create summary sheet with key metrics"""
        summary_data = {
            'Metric': [
                'Total Beds',
                'Average Occupancy Rate',
                'Break-even Occupancy',
                'Average Monthly Revenue',
                'Average Monthly Expenses',
                'Average Monthly Profit',
                'Annual Revenue (Year 1)',
                'Annual Profit (Year 1)',
                'Profit Margin'
            ],
            'Value': [
                self.total_beds,
                f"{projections_df['Occupancy_Rate'].mean():.1%}",
                f"{self.calculate_break_even():.1%}",
                f"${projections_df['Revenue'].mean():,.0f}",
                f"${projections_df['Expenses'].mean():,.0f}",
                f"${projections_df['Net_Income'].mean():,.0f}",
                f"${projections_df[projections_df['Year'] == self.start_date.year]['Revenue'].sum():,.0f}",
                f"${projections_df[projections_df['Year'] == self.start_date.year]['Net_Income'].sum():,.0f}",
                f"{(projections_df['Net_Income'].sum() / projections_df['Revenue'].sum()):.1%}"
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_assumptions_sheet(self, writer):
        """Create assumptions sheet"""
        assumptions_data = []
        
        # Room types
        assumptions_data.append(['Room Types', '', ''])
        assumptions_data.append(['Type', 'Beds', 'Daily Rate'])
        for room_type, details in self.room_types.items():
            assumptions_data.append([room_type, details['beds'], f"${details['rate']}"])
        
        assumptions_data.append(['', '', ''])
        
        # Occupancy rates
        assumptions_data.append(['Occupancy Rates', '', ''])
        for season, rate in self.assumptions['occupancy_rate'].items():
            assumptions_data.append([season, f"{rate:.0%}", ''])
        
        assumptions_data.append(['', '', ''])
        
        # Operating expenses
        assumptions_data.append(['Monthly Operating Expenses', '', ''])
        for expense, amount in self.assumptions['operating_expenses'].items():
            assumptions_data.append([expense, f"${amount:,}", ''])
        
        df_assumptions = pd.DataFrame(assumptions_data, columns=['Category', 'Value', 'Notes'])
        df_assumptions.to_excel(writer, sheet_name='Assumptions', index=False)
    
    def _create_dashboard_sheet(self, writer, projections_df):
        """Create dashboard with charts"""
        # Prepare data for charts
        monthly_summary = projections_df.groupby('Month_Name').agg({
            'Revenue': 'mean',
            'Expenses': 'mean',
            'Net_Income': 'mean'
        }).round(0)
        
        yearly_summary = projections_df.groupby('Year').agg({
            'Revenue': 'sum',
            'Expenses': 'sum',
            'Net_Income': 'sum'
        }).round(0)
        
        # Write data
        monthly_summary.to_excel(writer, sheet_name='Dashboard', startrow=1, startcol=0)
        yearly_summary.to_excel(writer, sheet_name='Dashboard', startrow=1, startcol=5)
    
    def _format_excel_sheets(self, writer):
        """Apply formatting to Excel sheets"""
        workbook = writer.book
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        currency_format = '#,##0'
        percent_format = '0.0%'
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Apply header formatting
            for cell in worksheet[1]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


if __name__ == "__main__":
    # Create financial model
    model = HostelFinancialModel(hostel_name="Hostel Diary")
    model.create_excel_model('hostel_diary_financial_model.xlsx')
    print("Financial model generation complete!")
